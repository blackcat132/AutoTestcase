# -*- coding: utf-8 -*-
"""
作者: holmes
日期: 2021年 07月 2021/7/29日 16:06
"""
##_________  专门处理文件__________ 封装
import json
from openpyxl import load_workbook
import openpyxl

def parse_json_file(filepath):
    data = json.load(open("D:/AutoApi/autotest_xudada_07/abtest/data/topics.json", mode="r", encoding="utf8"))  # 加载权限
    test_data = data['test_data']
    return test_data

def parse_excel_file(filepath,sheetname):
    web = openpyxl.load_workbook('D:/AutoApi/autotest_xudada_07/abtest/data/datas.xlsx')
    print(web.worksheets)
    ws = web[sheetname]
    # print(ws['B1'].value)
    print('查看行数:', len(tuple(ws.rows)))  # rows 行
    #### 实体 行3  列6  去首行 首列
    rows = ws.max_row
    columns = ws.max_column
    print("cc", columns)

    test_data = []
    for x in range(2, len(tuple(ws.rows)) + 1):  # 循环 行
        testcase_data_01 = []
        for y in range(2, columns + 1):  # 打印 2- 7行   注: 包头不包尾 +1
            testcase_data_01.append(ws.cell(row=x, column=y).value)
            print(ws.cell(row=x, column=y).value)
        test_data.append(testcase_data_01)

    return test_data
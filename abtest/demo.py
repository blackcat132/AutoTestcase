# -*- coding: utf-8 -*-
"""
作者: holmes
日期: 2021年 07月 2021/7/29日 10:23
./abtest/data/topics.json
D:/AutoApi/autotest_xudada_07/abtest/data/topics.json
"""
#D:/AutoApi/autotest_xudada_07/abtest/data/topics.json

from openpyxl import load_workbook
import openpyxl



wb = openpyxl.load_workbook('D:/AutoApi/autotest_xudada_07/abtest/data/datas.xlsx')
# print(web.worksheets)
ws = wb['topics']
# print(ws['D1'].value)
print('查看行数:',len(tuple(ws.rows)))   # rows 行 6行
#### 实体 行3  列6  去首行 首列
rows=ws.max_row
columns = ws.max_column
print("cc",columns)           ##cc是7



test_data =[]
for x in range(2, len(tuple(ws.rows)) + 1):  #循环 行
    testcase_data_01 = []
    for y in range(2,columns+1):  # 打印 2- 7行
        testcase_data_01.append(ws.cell(row=x, column=y).value)
        print(ws.cell(row=x, column=y).value)
#




# from openpyxl import Workbook
# wb = Workbook()
#
# # grab the active worksheet
# ws = wb.active
#
# # Data can be assigned directly to cells
# ws['A1'] = 42
#
# # Rows can also be appended
# ws.append([1, 2, 3])
#
# # Python types will automatically be converted
# import datetime
# ws['A2'] = datetime.datetime.now()
#
# # Save the file
# wb.save("sample.xlsx")



# import json
#              # 绝对路径
# data = json.load(open("D:/AutoApi/autotest_xudada_07/abtest/data/topics.json", mode="r", encoding="utf8"))  # 加载权限
# print(type(data))
# print(data)
#


# with open('./datas/topics.json','r',encoding='utf8')as fp:
#     json_data = json.load(fp)
#     print('这是文件中的json数据：',json_data)
#     print('这是读取到文件数据的数据类型：', type(json_data


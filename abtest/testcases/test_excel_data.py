# -*- coding: utf-8 -*-
"""
作者: holmes
日期: 2021年 07月 2021/7/29日 15:45
"""
'''
执行pytest:  pytest abtest\testcases\test_excel_data.py

pytest abtest\testcases\test_excel_data.py -v

'''
from openpyxl import load_workbook
import openpyxl
import pytest
import requests
import json
from abtest.business.file_utils import parse_excel_file

# 把 abtest.business.file_utils下的    parse_excel_file


# web = openpyxl.load_workbook('D:/AutoApi/autotest_xudada_07/abtest/data/datas.xlsx')
# # print(web.worksheets)
# ws = web['topics']
# # print(ws['B1'].value)
# print('查看行数:',len(tuple(ws.rows)))   # rows 行
# #### 实体 行3  列6  去首行 首列
# rows=ws.max_row
# columns = ws.max_column
# print("cc",columns)
#
# test_data = []
# for x in range(2, len(tuple(ws.rows)) + 1):  # 循环 行
#     testcase_data_01 = []
#     for y in range(2, columns + 1):  # 打印 2- 7行
#         testcase_data_01.append(ws.cell(row=x, column=y).value)
#         print(ws.cell(row=x, column=y).value)
#     test_data.append(testcase_data_01)

test_data = parse_excel_file('D:/AutoApi/autotest_xudada_07/abtest/data/datas.xlsx','topics')



# headers=header

@pytest.mark.parametrize('url,method,topic_params,Rstatus_code,r_msg',test_data)
def test_create_topic(url,method,topic_params,Rstatus_code,r_msg):
    if method =='post':
        # headers ={"ticket":"mobplus_a8ec0d1f93fb4404a10994df2c4c14e9nO0Y1j89a2500477951",
        #           "Content-Type":"application/json"}
        res_post=requests.post(url,data=json.loads(topic_params))   # headers=headers
        assert res_post.status_code ==Rstatus_code
        assert res_post.json()== json.loads(r_msg)
        # assert res_post.json()['msg'] == "成功"

    elif method =='get':
        res_get=requests.get(url)
        print(res_get.json())
        assert res_get.status_code ==Rstatus_code
        assert res_get.json()['success'] == True
        # assert res_get.json()== json.loads(r_msg)
    else:
        print("无该方法")




